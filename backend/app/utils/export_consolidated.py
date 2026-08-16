"""
Export builders for the Consolidated PPMP report.

Excel: openpyxl (already a light dependency; add to requirements.txt if
not already present: `openpyxl`)
PDF: reportlab (add `reportlab` to requirements.txt if not already present)

Both take the same ConsolidatedPPMPResponse the page renders from, so the
export always matches exactly what the admin sees on screen.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.schemas.consolidation import ConsolidatedPPMPResponse

PESO = "\u20b1"


def _fmt(n: float) -> str:
    return f"{PESO}{n:,.2f}"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def export_to_excel(data: ConsolidatedPPMPResponse) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated PPMP"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    office_fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")  # matches yellow office row on screen
    subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    ws.append([f"Consolidated PPMP — {data.fee_category}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Fiscal Year: {data.fiscal_year}    PPMP Type: {data.ppmp_type}"])
    ws.append([])

    columns = ["Office", "Project", "Entry / Code", "Item", "Qty", "Unit", "Unit Price", "Amount"]
    header_row = ws.max_row + 1
    ws.append(columns)
    for cell in ws[header_row]:
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for office in data.offices:
        office_row = ws.max_row + 1
        ws.append([office.office_name])
        for cell in ws[office_row]:
            cell.fill = office_fill
            cell.font = bold

        for project in office.projects:
            first_entry_row = True
            for entry in project.entries:
                for item in entry.items:
                    ws.append([
                        "",
                        project.project_label if first_entry_row else "",
                        entry.category_description,
                        item.item_name,
                        item.quantity,
                        item.unit,
                        item.unit_price,
                        item.total_cost,
                    ])
                    first_entry_row = False
                entry_row = ws.max_row + 1
                ws.append(["", "", "", "Subtotal", "", "", "", entry.entry_subtotal])
                for cell in ws[entry_row]:
                    cell.font = bold

            proj_row = ws.max_row + 1
            ws.append(["", "Project Subtotal", "", "", "", "", "", project.project_subtotal])
            for cell in ws[proj_row]:
                cell.fill = subtotal_fill
                cell.font = bold

        office_total_row = ws.max_row + 1
        ws.append([f"Office Total — {office.office_name}", "", "", "", "", "", "", office.office_total])
        for cell in ws[office_total_row]:
            cell.font = white_bold
            cell.fill = header_fill

        ws.append([])

    grand_row = ws.max_row + 1
    ws.append(["GRAND TOTAL", "", "", "", "", "", "", data.grand_total])
    for cell in ws[grand_row]:
        cell.font = Font(bold=True, size=12)

    for col, width in zip("ABCDEFGH", [22, 26, 8, 30, 8, 8, 14, 16]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def export_to_pdf(data: ConsolidatedPPMPResponse) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Consolidated PPMP Report \u2014 {data.fee_category}", styles["Title"]))
    elements.append(Paragraph(
        f"Fiscal Year: {data.fiscal_year} &nbsp;&nbsp; PPMP Type: {data.ppmp_type}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 8))

    for office in data.offices:
        # Yellow office banner, matching the on-screen office row
        office_banner = Table(
            [[Paragraph(f"<b>Office: {office.office_name}</b>", styles["Normal"]),
              Paragraph(f"<b>Office Total: {_fmt(office.office_total)}</b>", styles["Normal"])]],
            colWidths=[110 * mm, 55 * mm],
        )
        office_banner.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFEB99")),
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(office_banner)
        elements.append(Spacer(1, 4))

        for project in office.projects:
            elements.append(Paragraph(f"Project: {project.project_label}", styles["Heading3"]))

            for entry in project.entries:
                elements.append(Paragraph(f"<b>{entry.description}</b> ({entry.category_description})", styles["Normal"]))
                rows = [["Item", "Qty", "Unit", "Unit Price", "Amount"]]
                for item in entry.items:
                    rows.append([
                        item.item_name, f"{item.quantity:g}", item.unit,
                        _fmt(item.unit_price), _fmt(item.total_cost),
                    ])
                rows.append(["", "", "", "Subtotal", _fmt(entry.entry_subtotal)])

                t = Table(rows, colWidths=[70 * mm, 15 * mm, 20 * mm, 30 * mm, 30 * mm])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F7FA")]),
                    ("SPAN", (0, -1), (2, -1)),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 4))

            elements.append(Paragraph(f"<b>Project Subtotal: {_fmt(project.project_subtotal)}</b>", styles["Normal"]))
            elements.append(Spacer(1, 6))

        elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>GRAND TOTAL: {_fmt(data.grand_total)}</b>", styles["Title"]))

    doc.build(elements)
    buf.seek(0)
    return buf