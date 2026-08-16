"""
Export builders for the Consolidated APP report.

Excel: openpyxl
PDF: reportlab

Both take the same ConsolidatedAPPResponse the page renders from.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.schemas.app_consolidation import ConsolidatedAPPResponse

PESO = "\u20b1"


def _fmt(n: float) -> str:
    return f"{PESO}{n:,.2f}"


APP_COLUMN_HEADERS = [
    "Project Title",
    "General Description of the Project",
    "Mode of Procurement",
    "Early Procurement Activity? (Yes/No)",
    "Criteria for Bid Evaluation",
    "Start of Procurement Activity",
    "End of Procurement Activity",
    "Source of Fund",
    "Estimated Budget / ABC (PhP)",
    "Procurement Strategy or Tools",
    "Remarks",
]


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def export_to_excel(data: ConsolidatedAPPResponse) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated APP"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    cat_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    ws.append([f"Consolidated APP — {data.fee_category}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Fiscal Year: {data.fiscal_year}    Version: {data.app_version_type.upper()}"])
    ws.append([f"Number of PPMPs: {data.ppmp_count}"])
    ws.append([])

    columns = ["Project Title", "General Description", "Mode of Procurement",
               "EPA", "Bid Evaluation", "Start", "End", "Source of Fund",
               "Estimated Budget", "Procurement Strategy", "Remarks"]
    header_row = ws.max_row + 1
    ws.append(columns)
    for cell in ws[header_row]:
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for cat in data.categories:
        cat_row = ws.max_row + 1
        ws.append([cat.label])
        for cell in ws[cat_row]:
            cell.fill = cat_fill
            cell.font = bold

        for row in cat.rows:
            ws.append([
                row.project_title,
                row.general_description,
                row.procurement_mode,
                row.early_procurement,
                row.bid_evaluation,
                row.start_activity,
                row.end_activity,
                row.source_of_funds,
                row.estimated_budget,
                ", ".join(row.procurement_strategy) if row.procurement_strategy else "",
                row.remarks,
            ])

        sub_row = ws.max_row + 1
        ws.append(["", "", "", "", "", "", "", "Subtotal", cat.subtotal, "", ""])
        for cell in ws[sub_row]:
            cell.font = bold

    grand_row = ws.max_row + 1
    ws.append(["", "", "", "", "", "", "", "GRAND TOTAL", data.grand_total, "", ""])
    for cell in ws[grand_row]:
        cell.font = Font(bold=True, size=12)

    for col, width in zip("ABCDEFGHIJK", [22, 30, 14, 8, 10, 10, 10, 12, 16, 20, 14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def export_to_pdf(data: ConsolidatedAPPResponse) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Consolidated Annual Procurement Plan — {data.fee_category}", styles["Title"]))
    elements.append(Paragraph(
        f"Fiscal Year: {data.fiscal_year} &nbsp;&nbsp; Version: {data.app_version_type.upper()} &nbsp;&nbsp; PPMPs: {data.ppmp_count}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 8))

    for cat in data.categories:
        elements.append(Paragraph(f"<b>{cat.label}</b>", styles["Heading3"]))

        rows = [["Project Title", "General Description", "Mode", "EPA", "Start", "End", "Budget", "Strategy"]]
        for row in cat.rows:
            rows.append([
                row.project_title,
                row.general_description[:60] + ("..." if len(row.general_description) > 60 else ""),
                row.procurement_mode,
                row.early_procurement,
                row.start_activity,
                row.end_activity,
                _fmt(row.estimated_budget),
                ", ".join(row.procurement_strategy[:2]) if row.procurement_strategy else "",
            ])
        rows.append(["", "", "", "", "", "Subtotal", _fmt(cat.subtotal), ""])

        col_widths = [35 * mm, 45 * mm, 20 * mm, 12 * mm, 18 * mm, 18 * mm, 22 * mm, 30 * mm]
        t = Table(rows, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F7FA")]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 8))

    elements.append(Paragraph(f"<b>GRAND TOTAL: {_fmt(data.grand_total)}</b>", styles["Title"]))

    doc.build(elements)
    buf.seek(0)
    return buf
