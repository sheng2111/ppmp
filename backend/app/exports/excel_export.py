import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

MONTHS = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
MONTH_LABELS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

Q1_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green
Q2_FILL = PatternFill("solid", fgColor="FFEB9C")   # light yellow  
Q3_FILL = PatternFill("solid", fgColor="FFC7CE")   # light red/pink
TOTAL_FILL = PatternFill("solid", fgColor="FFFF00") # yellow for sub-total
HEADER_FILL = PatternFill("solid", fgColor="1A56A4") # primary blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
BOLD_FONT = Font(bold=True, size=9)
NORMAL_FONT = Font(size=9)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def currency_fmt():
    return '#,##0.00'

def generate_ppmp_excel(ppmp_data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "PPMP"
    ws.page_setup.fitToWidth = 1
    ws.print_options.horizontalCentered = True

    header = ppmp_data["header"]
    items = ppmp_data["items"]

    # ---- Document Header ----
    ws.merge_cells('A1:B1')
    ws['A1'] = f"END-USER/UNIT: {header.get('end_user_unit', '')}"
    ws['A1'].font = NORMAL_FONT

    ws.merge_cells('G1:J1')
    ws['G1'] = "PROJECT PROCUREMENT MANAGEMENT PLAN (PPMP)"
    ws['G1'].font = Font(bold=True, size=11)
    ws['G1'].alignment = CENTER

    ws['N1'] = f"Date: {header.get('date', '')}"
    ws['N1'].font = NORMAL_FONT

    ws.merge_cells('A2:B2')
    ws['A2'] = f"Charged to: {header.get('charged_to', '')}"
    ws['A2'].font = NORMAL_FONT

    ws['N2'] = f"Revision #: {header.get('revision', '')}"
    ws['N2'].font = NORMAL_FONT

    ws.merge_cells('A3:C3')
    ws['A3'] = f"Projects, Activities and Programs (PAPs): {header.get('pap', '')}"
    ws['A3'].font = NORMAL_FONT

    # ---- Column Headers (Row 5-6) ----
    col_headers = [
        ("A", 8, "CODE"),
        ("B", 22, "GENERAL DESCRIPTION"),
        ("C", 8, "Unit of Issue"),
        ("D", 7, "QUANTITY /SIZE"),
        ("E", 10, "Unit Cost"),
        ("F", 12, "Total cost"),
        ("G", 10, "Mode of Procurement"),
    ]

    for col_letter, width, title in col_headers:
        ws.column_dimensions[col_letter].width = width
        cell = ws[f"{col_letter}5"]
        cell.value = title
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border()
        ws.merge_cells(f"{col_letter}5:{col_letter}6")

    # Monthly schedule columns
    month_start_col = 8  # Column H
    month_fills = [Q1_FILL]*3 + [Q1_FILL]*3 + [Q2_FILL]*3 + [Q2_FILL]*3  # simplified

    # Q1 total, Q2 total, Q3 total header merges
    # Jan-Mar = H-J (cols 8-10), Total Q1 = col 11
    # Apr-Jun = L-N (cols 12-14), Total Q2 = col 15
    # Jul-Sep = P-R (cols 16-18), Total Q3 = col 19
    # Oct-Dec = T-V (cols 20-22)

    month_col_map = {}
    col = month_start_col

    quarters = [
        (["jan","feb","mar"], "TOTAL Q1", Q1_FILL),
        (["apr","may","jun"], "TOTAL Q2", Q2_FILL),
        (["jul","aug","sep"], "TOTAL Q3", Q3_FILL),
        (["oct","nov","dec"], None, None),
    ]

    # Header row for "SCHEDULE/MILESTONE OF ACTIVITIES"
    sched_start = get_column_letter(month_start_col)
    ws[f"{sched_start}5"] = "SCHEDULE/MILESTONE OF ACTIVITIES"
    ws[f"{sched_start}5"].font = HEADER_FONT
    ws[f"{sched_start}5"].fill = HEADER_FILL
    ws[f"{sched_start}5"].alignment = CENTER

    # We'll place months with Qty/Amt pairs
    for qidx, (months_in_q, q_label, q_fill) in enumerate(quarters):
        for midx, m in enumerate(months_in_q):
            qty_col = col
            amt_col = col + 1
            qty_letter = get_column_letter(qty_col)
            amt_letter = get_column_letter(amt_col)
            month_col_map[m] = (qty_col, amt_col)

            # Month label spanning Qty+Amt
            ws.merge_cells(f"{qty_letter}6:{amt_letter}6")
            cell = ws[f"{qty_letter}6"]
            cell.value = MONTH_LABELS[MONTHS.index(m)]
            cell.font = BOLD_FONT
            fill = [Q1_FILL, Q1_FILL, Q1_FILL, Q2_FILL, Q2_FILL, Q2_FILL, Q3_FILL, Q3_FILL, Q3_FILL, None, None, None][MONTHS.index(m)]
            if fill: cell.fill = fill
            cell.alignment = CENTER
            cell.border = thin_border()

            ws.column_dimensions[qty_letter].width = 6
            ws.column_dimensions[amt_letter].width = 10

            col += 2

        # Quarter total columns (Qty + Amt)
        if q_label:
            qt_qty = col
            qt_amt = col + 1
            qt_qty_letter = get_column_letter(qt_qty)
            qt_amt_letter = get_column_letter(qt_amt)
            ws.merge_cells(f"{qt_qty_letter}6:{qt_amt_letter}6")
            cell = ws[f"{qt_qty_letter}6"]
            cell.value = q_label
            cell.font = BOLD_FONT
            if q_fill: cell.fill = q_fill
            cell.alignment = CENTER
            cell.border = thin_border()
            ws.column_dimensions[qt_qty_letter].width = 7
            ws.column_dimensions[qt_amt_letter].width = 12
            col += 2

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 25

    # ---- Data Rows ----
    row = 7
    categories = {}
    for item in items:
        cat = item.get('category') or 'General'
        categories.setdefault(cat, []).append(item)

    grand_total = 0.0

    for cat_name, cat_items in categories.items():
        # Category header row
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = cat_name
        ws[f"A{row}"].font = Font(bold=True, size=9, italic=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="D9E1F2")
        ws[f"A{row}"].alignment = LEFT
        row += 1

        cat_total = 0.0
        for item in cat_items:
            schedule = item.get('schedule', {})
            total = item.get('total_cost', 0)
            cat_total += total
            grand_total += total

            ws[f"A{row}"] = item.get('code', '')
            ws[f"B{row}"] = item.get('general_description', '')
            ws[f"C{row}"] = item.get('unit_of_issue', '')
            ws[f"D{row}"] = item.get('quantity', 0)
            ws[f"E{row}"] = item.get('unit_cost', 0)
            ws[f"F{row}"] = total
            ws[f"G{row}"] = item.get('mode_of_procurement', '')

            ws[f"E{row}"].number_format = currency_fmt()
            ws[f"F{row}"].number_format = currency_fmt()

            for col_letter in ['A','B','C','D','E','F','G']:
                ws[f"{col_letter}{row}"].font = NORMAL_FONT
                ws[f"{col_letter}{row}"].border = thin_border()
                ws[f"{col_letter}{row}"].alignment = LEFT if col_letter in ['B','G'] else CENTER

            # Fill monthly schedule
            col = month_start_col
            for qidx, (months_in_q, q_label, q_fill) in enumerate(quarters):
                q_amt = 0.0
                for m in months_in_q:
                    sched = schedule.get(m, {})
                    qty_val = sched.get('qty') or 0
                    amt_val = sched.get('amount') or 0
                    qty_col, amt_col = month_col_map[m]
                    qty_letter = get_column_letter(qty_col)
                    amt_letter = get_column_letter(amt_col)
                    ws[f"{qty_letter}{row}"] = qty_val if qty_val else '-'
                    ws[f"{amt_letter}{row}"] = amt_val if amt_val else '-'
                    ws[f"{qty_letter}{row}"].font = NORMAL_FONT
                    ws[f"{amt_letter}{row}"].font = NORMAL_FONT
                    ws[f"{qty_letter}{row}"].alignment = CENTER
                    ws[f"{amt_letter}{row}"].alignment = CENTER
                    ws[f"{qty_letter}{row}"].border = thin_border()
                    ws[f"{amt_letter}{row}"].border = thin_border()
                    if isinstance(amt_val, (int, float)) and amt_val:
                        ws[f"{amt_letter}{row}"].number_format = currency_fmt()
                    q_amt += (amt_val or 0)
                    col += 2

                if q_label:
                    qt_qty_letter = get_column_letter(col)
                    qt_amt_letter = get_column_letter(col + 1)
                    ws[f"{qt_qty_letter}{row}"] = '-'
                    ws[f"{qt_amt_letter}{row}"] = q_amt if q_amt else '-'
                    ws[f"{qt_qty_letter}{row}"].font = NORMAL_FONT
                    ws[f"{qt_amt_letter}{row}"].font = BOLD_FONT
                    ws[f"{qt_qty_letter}{row}"].alignment = CENTER
                    ws[f"{qt_amt_letter}{row}"].alignment = CENTER
                    ws[f"{qt_qty_letter}{row}"].border = thin_border()
                    ws[f"{qt_amt_letter}{row}"].border = thin_border()
                    if q_fill:
                        ws[f"{qt_amt_letter}{row}"].fill = q_fill
                    if isinstance(q_amt, (int, float)) and q_amt:
                        ws[f"{qt_amt_letter}{row}"].number_format = currency_fmt()
                    col += 2

            row += 1

        # Sub-total row
        ws[f"D{row}"] = "Sub-total"
        ws[f"D{row}"].font = BOLD_FONT
        ws[f"D{row}"].alignment = Alignment(horizontal='right')
        ws[f"F{row}"] = cat_total
        ws[f"F{row}"].font = BOLD_FONT
        ws[f"F{row}"].fill = TOTAL_FILL
        ws[f"F{row}"].number_format = currency_fmt()
        ws[f"F{row}"].border = thin_border()
        row += 1

    # Grand total
    ws[f"E{row}"] = "GRAND TOTAL"
    ws[f"E{row}"].font = Font(bold=True, size=10)
    ws[f"F{row}"] = grand_total
    ws[f"F{row}"].font = Font(bold=True, size=10)
    ws[f"F{row}"].fill = TOTAL_FILL
    ws[f"F{row}"].number_format = currency_fmt()
    ws[f"F{row}"].border = thin_border()

    # Freeze header rows
    ws.freeze_panes = "A7"

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
